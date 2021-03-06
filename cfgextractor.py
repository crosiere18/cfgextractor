#!/usr/bin/python3

from openpyxl import Workbook
import sys

def output(xaxis, yaxis, data):
    wb = Workbook()

    ws = wb.active
    ws.title = 'Main'

    xlen = len(xaxis) + 2
    ylen = len(yaxis) + 2

    for col in range(2, xlen):
        ws.cell(row = 1, column = col).value = xaxis[col-2]
        
    for row in range(2, ylen):
        ws.cell(row = row, column = 1).value = yaxis[row-2]
    
    for cidx, col in enumerate(data):
        for ridx, cdata in enumerate(col):
            ws.cell(row=ridx+2,column=cidx+2).value = cdata


    wb.save('output.xlsx')

def readin(files):
    ret = {}
    chunk = {}
    chunknum = 0 
    for fname in files:
        with open(fname, 'r') as cfgfile:
            for line in cfgfile:
                line = line.strip()
            
                if not line or line.startswith('#') or line.startswith('define'): continue
            
                if line.startswith('}'):
                    ret[chunknum] = chunk
                    chunknum += 1
                    chunk = {}
                    continue 

                bits = line.split(None, 1)

                if len(bits) > 1:
                    k, v = bits
                    chunk[k] = v

    return ret

def dechunkify(delim, axis, chunkmap, field):
    for chunk in chunkmap:
        if field in chunkmap[chunk]:
            values = chunkmap[chunk][field].split(delim)
            for item in values:
                item = item.strip()
                if item and item not in axis:
                    axis.append(item)

def start(filescfg):
    xaxis = []
    yaxis = []

    delim = ','
    
    infiles = open(filescfg, 'r')
    xfiles = infiles.readline().strip().split(delim)
    yfiles = infiles.readline().strip().split(delim)
    xfield = infiles.readline().strip()
    yfield = infiles.readline().strip()
    pivot = infiles.readline().strip()
    ydata = infiles.readline().strip().split(delim)

    xchunkmap = readin(xfiles)
    ychunkmap = readin(yfiles) 
    
    dechunkify(delim, xaxis, xchunkmap, xfield)

    dechunkify(delim, yaxis, ychunkmap, yfield)

    outdata = [['']*len(yaxis) for i in range(len(xaxis))]
            
    for chunk in ychunkmap:
        if pivot in ychunkmap[chunk] and yfield in ychunkmap[chunk]:
            youtdata = ''
            for field in ychunkmap[chunk]:
                if field in ydata:
                    youtdata = youtdata + field + ": " + ychunkmap[chunk][field] + '\n'

            for xidx, x in enumerate(xaxis):
                for yidx, y in enumerate(yaxis):
                    if x in ychunkmap[chunk][pivot] and y in ychunkmap[chunk][yfield]:
                        outdata[xidx][yidx] = youtdata
                    


    infiles.close()
    output(xaxis, yaxis, outdata)

if __name__ == "__main__":
    failstring = ("requires 6 arguments in one file: list of input files for x axis (with relative paths"
                "), list of input files for y axis, x axis field, y axis field, pivot/relation field, an"
                "d data required from y files. These should be listed in an input file and handed to the"
                " program ('ex. ./cfgextractor.py infiles.cfg'")

    if len(sys.argv) != 2:
        print(sys.argv[0] + failstring)
    else:
        start(sys.argv[1])   
